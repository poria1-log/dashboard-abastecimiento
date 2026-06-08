"""
Actualizador automático del Dashboard Central Oeste.
Corre vía GitHub Actions todos los días a las 8 AM Argentina.
Los tokens se leen de las variables de entorno (secretos de GitHub).
"""

import os
import re
import sys
import requests
import openpyxl
from io import BytesIO
from datetime import datetime
from github import Github, Auth

# ── TOKENS (desde secretos de GitHub) ────────────────────
KIKKER_TOKEN = os.environ.get("KIKKER_TOKEN", "")
GH_TOKEN     = os.environ.get("GH_TOKEN_PAT", "")

if not KIKKER_TOKEN:
    print("❌ KIKKER_TOKEN no configurado en los secretos de GitHub.")
    sys.exit(1)
if not GH_TOKEN:
    print("❌ GH_TOKEN_PAT no configurado en los secretos de GitHub.")
    sys.exit(1)

# ── CONFIGURACIÓN ─────────────────────────────────────────
GITHUB_USER = "poria1-log"
GITHUB_REPO = "dashboard-abastecimiento"
GITHUB_FILE = "dashboard-proveedores (27).html"

TODOS_LOS_MESES = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SET","OCT","NOV","DIC"]
MESES_ACTIVOS   = []

KIKKER_ENDPOINTS = [
    # MASIVOS
    { "dep":"Cuidado Bucal",              "family":"masivos",    "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=7915bf1c-6f45-4d43-b0ed-419db7ab4df0&ssd=coverage_days&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Cuidado de la Piel",         "family":"masivos",    "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=cb353838-67a9-4bb5-add7-8462594f77cb&ssd=coverage_days&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Cuidado del Bebé y la Mamá", "family":"masivos",    "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=342bbc63-8639-4c46-8e6d-0d79598e7cce&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Cuidado del Cabello",        "family":"masivos",    "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=1f66831e-7f6e-4cd0-a162-c05c8eafb891&ssd=coverage_days&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Cuidado Personal",           "family":"masivos",    "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=5b0d4077-6ae7-4305-a72f-42f6ad87b090&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Limpieza y Hogar",           "family":"masivos",    "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=0ebcf548-f271-4590-b2a2-3dc266020040&ssd=coverage_days&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    # ETICOS
    { "dep":"Accesorios",                 "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=0ddcae88-7ec3-4654-a961-a06a48b7acb2&ssd=coverage_days&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Electrosalud",               "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=47e0cd5d-5ef8-4ee2-a4aa-c1e4c4ab9415&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"In Vitro",                   "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=e2022c92-69ab-4691-8ee0-f8a46746008b&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Medicamentos",               "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=d932d853-b493-4ffd-88e4-488d2f98c93e&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Nutrición",                  "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=f99e0402-31d4-4278-9e9a-4564e9cf17da&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Parafarmacia",               "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=84d7be7f-c675-4225-8799-bf055b319f01&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Suplementos",                "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=e116dc00-43c1-4a66-8466-786ee2f9237f&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Venta Libre",                "family":"eticos",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=ab8b290a-eb6b-484d-a4c3-de810baea53c&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    # SELECTIVOS
    { "dep":"Accesorios",                 "family":"selectivos", "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&ssd=coverage_days&dop=1&se=b8059650-b690-4fe7-a6eb-7d08cf973ca3&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Belleza Profesional",        "family":"selectivos", "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=137e912e-d420-47a5-b477-ec0a8de6ee25&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Cosmética",                  "family":"selectivos", "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=4ff34edd-42bc-44c3-bf4c-28a4c8042a57&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Dermocosmética",             "family":"selectivos", "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=790725e2-b18b-4966-8f98-c4c4581d742a&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Fragancias",                 "family":"selectivos", "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&se=16ab0187-2650-4570-9923-55b282be7204&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    # MARKET
    { "dep":"Almacén",                    "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&ssd=coverage_days&dop=1&os=false&se=f0ce21bd-f74e-4133-98b8-cdcee0960dad&freq=monthly&cs=&pa=&seg=trade_in&seg2=actual_sales&locale=es-AR" },
    { "dep":"Bazar",                      "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=15ff5215-56fa-47b8-a1d4-fe218ee0d6ab&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Bebidas",                    "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=cdba99d4-da61-45db-b489-891633068d97&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Comestibles",                "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=24f40325-99b9-4807-9f12-41b951b03e9d&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Frescos",                    "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=5dffc294-ca1d-4ec1-bb46-9f7a99ff4713&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Galletitas",                 "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=3b650370-424f-41dc-bb4a-c9f7a748fe55&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Golosinas",                  "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=6a595115-3528-4cf9-a3fb-5ac7c9204219&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Hogar",                      "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=080ebfa1-023c-48ad-bb68-84519a3ff46e&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Juguetería",                 "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=9be764f2-ffc7-4627-88ba-e6793f64aa29&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Lácteos",                    "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=3aa7528d-0feb-4464-864c-53a4054be0bb&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
    { "dep":"Librería",                   "family":"market",     "url":"https://api-fco.kikker.com.br/api/dashboard/network_performance/export_excel?bg=b745953e-95df-4398-baa1-8e9fb40c349c&dc=0&os=false&se=349ffada-a8b5-4587-b86a-dba0311ab9b5&cs=&pa=&seg=trade_in&seg2=actual_sales&freq=monthly&locale=es-AR" },
]

# ── HELPERS ───────────────────────────────────────────────
def detectar_meses(content):
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    e26_col = 5
    for i, h in enumerate(header):
        if h and '2026' in str(h).upper() and 'ENTRADA' in str(h).upper():
            e26_col = i
            break
    meses = []
    for row in rows[1:]:
        if not row or not row[0]: continue
        mes = str(row[0]).strip().upper().replace('SEP','SET')
        if mes not in TODOS_LOS_MESES: continue
        val = row[e26_col] if len(row) > e26_col else None
        if val is not None and str(val).strip() not in ('', '0', 'None'):
            meses.append(mes)
    return [m for m in TODOS_LOS_MESES if m in meses]

def parse_num(val):
    if val is None: return 0
    val = str(val).replace('$','').replace('.','').replace(',','.').strip()
    try: return int(float(val))
    except: return 0

def procesar_excel(content, dep_name, family):
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    e25_col, v25_col, e26_col, v26_col = 3, 4, 5, 6
    for i, h in enumerate(header):
        if h is None: continue
        h = str(h).upper()
        if '2025' in h and h.startswith('2025 - ENTRADA'): e25_col = i
        if '2025' in h and h.startswith('2025 - VENTA'):   v25_col = i
        if '2026' in h and h.startswith('2026 - ENTRADA'): e26_col = i
        if '2026' in h and h.startswith('2026 - VENTA'):   v26_col = i
    e26=[0]*len(MESES_ACTIVOS); v26=[0]*len(MESES_ACTIVOS)
    e25=[0]*len(MESES_ACTIVOS); v25=[0]*len(MESES_ACTIVOS)
    for row in rows[1:]:
        if not row or not row[0]: continue
        mes = str(row[0]).strip().upper().replace('SEP','SET')
        if mes not in MESES_ACTIVOS: continue
        idx = MESES_ACTIVOS.index(mes)
        e26[idx] = parse_num(row[e26_col]) if len(row) > e26_col else 0
        v26[idx] = parse_num(row[v26_col]) if len(row) > v26_col else 0
        e25[idx] = parse_num(row[e25_col]) if len(row) > e25_col else 0
        v25[idx] = parse_num(row[v25_col]) if len(row) > v25_col else 0
    tag = {'masivos':'Masivos','eticos':'Éticos','selectivos':'Selectivos','market':'Market'}.get(family, family) + ' · Kikker'
    return {'name': dep_name, 'family': family, 'tag': tag,
            'e26': e26, 'v26': v26, 'e25': e25, 'v25': v25}

# ── MAIN ──────────────────────────────────────────────────
print(f"=== Actualizador Dashboard Central Oeste ===")
print(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
print(f"Departamentos configurados: {len(KIKKER_ENDPOINTS)}")

hdrs = {'Authorization': KIKKER_TOKEN}

# 1. Detectar meses
print("\n1. Detectando meses disponibles...")
r_probe = requests.get(KIKKER_ENDPOINTS[0]['url'], headers=hdrs, timeout=30)
if r_probe.status_code == 401:
    print("❌ Token de Kikker vencido. Actualizá el secreto KIKKER_TOKEN en GitHub.")
    sys.exit(1)
elif r_probe.status_code != 200:
    print(f"❌ Error al conectar con Kikker: HTTP {r_probe.status_code}")
    sys.exit(1)
MESES_ACTIVOS = detectar_meses(r_probe.content)
print(f"✅ Meses detectados: {MESES_ACTIVOS}")

# 2. Descargar y procesar
print("\n2. Descargando departamentos...")
providers = []
for ep in KIKKER_ENDPOINTS:
    r = requests.get(ep['url'], headers=hdrs, timeout=30)
    if r.status_code == 401:
        print("❌ Token de Kikker vencido.")
        sys.exit(1)
    elif r.status_code != 200:
        print(f"⚠️  Error HTTP {r.status_code} en {ep['dep']} — saltando.")
        continue
    p = procesar_excel(r.content, ep['dep'], ep['family'])
    total_e, total_v = sum(p['e26']), sum(p['v26'])
    ev = round(total_e/total_v*100) if total_v else 0
    print(f"   ✅ {ep['dep']} ({ep['family']}): E/V {ev}%")
    providers.append(p)

print(f"\n✅ {len(providers)} departamentos procesados")

# 3. Actualizar GitHub
print("\n3. Actualizando GitHub...")
lines = ['const PROVIDERS = [']
for p in providers:
    lines.append(f'  {{ name:"{p["name"]}", family:"{p["family"]}", tag:"{p["tag"]}", e26:{p["e26"]}, v26:{p["v26"]}, e25:{p["e25"]}, v25:{p["v25"]} }},')
lines.append('];')
nuevo_js = '\n'.join(lines)
nuevo_meses = f"const MESES = {MESES_ACTIVOS};"

g = Github(auth=Auth.Token(GH_TOKEN))
repo = g.get_repo(f"{GITHUB_USER}/{GITHUB_REPO}")
file = repo.get_contents(GITHUB_FILE)
html = file.decoded_content.decode('utf-8')

html = re.sub(r'const PROVIDERS = \[.*?\];', nuevo_js, html, flags=re.DOTALL)
html = re.sub(r"const MESES = \[.*?\];", nuevo_meses, html)

fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
meses_label = f"{MESES_ACTIVOS[0]}–{MESES_ACTIVOS[-1]} 2026" if MESES_ACTIVOS else "2026"
html = re.sub(
    r'id="info-banner-text">.*?<',
    f'id="info-banner-text">\U0001f7e2 &nbsp;Datos sincronizados desde <strong>Kikker</strong> · {len(providers)} departamentos · {meses_label} · Última actualización: {fecha}<',
    html, flags=re.DOTALL
)

repo.update_file(
    path=GITHUB_FILE,
    message=f"Auto-update · {fecha}",
    content=html.encode('utf-8'),
    sha=file.sha
)

print(f"✅ Dashboard actualizado — {len(providers)} departamentos · {meses_label}")
print(f"🌐 https://{GITHUB_USER}.github.io/{GITHUB_REPO}/{GITHUB_FILE.replace(' ', '%20')}")
